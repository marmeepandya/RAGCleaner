#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import sys
sys.path.insert(0, '/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI/venv/lib64/python3.12/site-packages')
sys.path.insert(0, '/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI/venv/lib/python3.12/site-packages')
sys.path.append('/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI')

import os, re, glob, time, random, threading
import torch
import numpy as np
import pandas as pd
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage
from sentence_transformers import CrossEncoder, util

random.seed(42); np.random.seed(42); torch.manual_seed(42)

print(f'PyTorch: {torch.__version__}')
print(f'CUDA: {torch.cuda.is_available()} -- {torch.cuda.get_device_name(0) if torch.cuda.is_available() else "CPU"}')

TARGET_ATTRIBUTES  = ['bus_type','model_number','model','read_speed_mb_s','write_speed_mb_s','height_mm','width_mm']
NUMERIC_ATTRIBUTES = {'read_speed_mb_s','write_speed_mb_s','height_mm','width_mm'}
TEXT_ATTRIBUTES    = {'bus_type','model_number','model'}
HF_CACHE           = '/home/ma/ma_ma/ma_mpandya/.cache/huggingface/hub'
EMBEDDINGS_DIR     = 'embeddings'
SKIP_FIELDS        = {'id','url','description','title_description','price','priceCurrency','cluster_id'}
TOP_N              = 20
TOP_K              = 5
CHECKPOINT_EVERY   = 5

DATA_DIR = 'normalized_products'
df1      = pd.read_json(f'{DATA_DIR}/dataset_1_normalized.json')
df2      = pd.read_json(f'{DATA_DIR}/dataset_2_normalized.json')
df3      = pd.read_json(f'{DATA_DIR}/dataset_3_normalized.json')
df4      = pd.read_json(f'{DATA_DIR}/dataset_4_normalized.json')
kb_full  = pd.concat([df2, df3, df4], ignore_index=True)
kb       = kb_full.copy()

assert os.path.exists('eval_set.csv'), 'Run exp_setup.ipynb first'
eval_df          = pd.read_csv('eval_set.csv')
query_indices    = pd.read_csv('query_indices.csv').iloc[:,0].tolist()
query_df         = df1.loc[query_indices].copy()
query_idx_to_pos = {idx: pos for pos, idx in enumerate(query_df.index)}

print(f'KB: {len(kb):,} rows | Query rows: {len(query_df)} | Eval tasks: {len(eval_df)}')

print('Loading BGE embeddings...')
bge_kb_embs    = torch.load(f'{EMBEDDINGS_DIR}/bge_kb.pt',    map_location='cpu')
bge_query_embs = torch.load(f'{EMBEDDINGS_DIR}/bge_query.pt', map_location='cpu')
print(f'  BGE KB: {bge_kb_embs.shape} | Query: {bge_query_embs.shape}')

CE_SNAP = glob.glob(f'{HF_CACHE}/models--cross-encoder--ms-marco-MiniLM-L-6-v2/snapshots/*/')
CE_PATH = CE_SNAP[0].rstrip('/') if CE_SNAP else 'cross-encoder/ms-marco-MiniLM-L-6-v2'
cross_encoder = CrossEncoder(CE_PATH)
print(f'CrossEncoder: {CE_PATH}')

def is_correct_standard(predicted, ground_truth, attribute):
    if not predicted or str(predicted).strip().lower() in {'','nan','none','unknown','null'}:
        return False
    if attribute in NUMERIC_ATTRIBUTES:
        try:
            p = float(str(predicted).replace(',','').strip())
            g = float(str(ground_truth).replace(',','').strip())
            return p == g
        except: return False
    p, g = str(predicted).lower().strip(), str(ground_truth).lower().strip()
    return p == g or p in g or g in p

def evaluate_ce(predicted, ground_truth, attribute):
    if predicted == 'UNKNOWN' or str(predicted).lower() in {'nan','none','null',''}:
        return 'wrong'
    if attribute in NUMERIC_ATTRIBUTES:
        try:
            p = float(str(predicted).replace(',','').strip())
            g = float(str(ground_truth).replace(',','').strip())
            if g == 0: return 'correct' if p == 0 else 'wrong'
            r = abs(p-g)/abs(g)
            return 'correct' if r == 0.0 else ('acceptable' if r <= 0.10 else 'wrong')
        except: return 'wrong'
    score = cross_encoder.predict([[ground_truth, predicted]])[0]
    return 'correct' if score > 2.0 else ('acceptable' if score > -1.0 else 'wrong')

def parse_response(text, attribute):
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith('VALUE:'):
            val = line.split(':',1)[1].strip().strip('"').strip("'")
            return 'UNKNOWN' if val.upper() in {'UNKNOWN','NONE','NAN','NULL',''} else val
    pat = rf'{attribute}\s*[:\u2192>\-]+\s*([^\s|]+)'
    m = re.search(pat, text, re.IGNORECASE)
    if m:
        val = m.group(1).strip().strip('"').strip("'").strip('[]')
        if val.upper() != 'UNKNOWN' and len(val)>2 and val.lower() not in {'none','nan','null','exact','value'}:
            return val
    if attribute in NUMERIC_ATTRIBUTES:
        nums = re.findall(r'\b\d+\.?\d*\b', text)
        if nums: return nums[0]
    cleaned = text.strip().strip('"').strip("'")
    if cleaned.upper().startswith('VALUE:'): cleaned = cleaned.split(':',1)[1].strip()
    if cleaned and len(cleaned)<80 and '\n' not in cleaned and cleaned.upper() not in {'UNKNOWN','NONE','NULL','NAN',''}:
        return cleaned
    return 'UNKNOWN'

def fix_prediction(pred):
    if isinstance(pred, str) and pred.strip().upper().startswith('VALUE:'):
        val = pred.strip().split(':',1)[1].strip()
        return 'UNKNOWN' if val.upper() in {'UNKNOWN','NONE','NULL','NAN',''} else val
    return pred

def evaluate_and_save(results_df, config_name, filepath):
    results_df['predicted']        = results_df['predicted'].apply(fix_prediction)
    results_df['unknown']          = results_df['predicted'] == 'UNKNOWN'
    results_df['correct_standard'] = results_df.apply(
        lambda r: is_correct_standard(r['predicted'], r['ground_truth'], r['attribute']), axis=1)
    results_df['ce_judgment']      = [
        evaluate_ce(r['predicted'], r['ground_truth'], r['attribute'])
        for _, r in results_df.iterrows()]
    results_df.to_csv(filepath, index=False)
    std = results_df['correct_standard'].mean()
    ce  = results_df['ce_judgment'].isin(['correct','acceptable']).mean()
    unk = results_df['unknown'].mean()
    print(f'\n{"="*60}\nRESULTS -- {config_name}\n{"="*60}')
    print(f'Standard accuracy:    {std:.3f} ({std*100:.1f}%)')
    print(f'CE eval (c+a):        {ce:.3f} ({ce*100:.1f}%)')
    print(f'UNKNOWN rate:         {unk:.3f} ({unk*100:.1f}%)')
    print(f'Total tasks:          {len(results_df)}')
    print('\nPer-attribute:')
    print(results_df.groupby('attribute').agg(
        n=('correct_standard','count'),
        std_acc=('correct_standard','mean'),
        ce_acc=('ce_judgment', lambda x: x.isin(['correct','acceptable']).mean()),
        unknown=('unknown','mean')
    ).round(3).to_string())
    print(f'\n+ Saved to {filepath}')
    return results_df

print('Eval functions ready.')

PROMPTS_OHNE = {
'bus_type': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
You MUST always provide a value. Pick the best matching reference product even if uncertain.
Step 1: Find the reference product that best matches the query product.
Step 2: Copy the bus_type value from that reference product exactly.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache - WD60EZAZ
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | bus_type: SATA III | model_number: WD60EZAZ
Best match: WD Blue 6TB WD60EZAZ -> bus_type: SATA III
VALUE:SATA III

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | bus_type: PCIe 3.0 x4 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> bus_type: PCIe 3.0 x4
VALUE:PCIe 3.0 x4

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> bus_type: [value from reference]
VALUE:""",

'model_number': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT generate or guess a model number.
Copy the EXACT model_number from the best matching reference product character by character.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

CRITICAL: model_numbers look like GV-N3080GAMING OC-10GD or CSSD-F960GBMP510.
Pay attention to every character -- GV-N166SOC-6GD and GV-N1660OC-6GD are DIFFERENT products.
If uncertain between similar SKUs, pick the one whose title most closely matches the query.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - SATA 6Gb/s 256MB Cache 3.5 Inch
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB Hard Drive -> model_number: WD60EZAZ
VALUE:WD60EZAZ

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD PCIe Gen3 x4 NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> model_number: CSSD-F960GBMP510
VALUE:CSSD-F960GBMP510

Example 3 (near-identical SKUs):
Query: Gigabyte GeForce GTX 1660 SUPER OC 6G graphics card
Reference products:
  - title: Gigabyte GTX 1660 Ti OC 6G | model_number: GV-N166TOC-6GD | brand: Gigabyte
  - title: Gigabyte GTX 1660 SUPER OC 6G | model_number: GV-N166SOC-6GD | brand: Gigabyte
Best match: GTX 1660 SUPER OC (not Ti) -> model_number: GV-N166SOC-6GD
VALUE:GV-N166SOC-6GD

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model_number: [exact value from reference]
VALUE:""",

'model': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
Copy the exact model name from the best matching reference product.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB -> model: WD Blue
VALUE:WD Blue

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 -> model: Force Series MP510
VALUE:Force Series MP510

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model: [value from reference]
VALUE:""",

'read_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact read_speed_mb_s number.
Return a number only. Do NOT return the write speed.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> read_speed_mb_s: 3480
VALUE:3480

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> read_speed_mb_s: [value from reference]
VALUE:""",

'write_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact write_speed_mb_s number.
Return a number only. Do NOT return the read speed.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> write_speed_mb_s: 3000
VALUE:3000

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> write_speed_mb_s: [value from reference]
VALUE:""",

'height_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact height_mm number.
Return a number only. Do NOT confuse height with width or length.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> height_mm: 46
VALUE:46

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> height_mm: [value from reference]
VALUE:""",

'width_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact width_mm number.
Return a number only. Do NOT confuse width with height or length.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> width_mm: 127
VALUE:127

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> width_mm: [value from reference]
VALUE:"""
}

PROMPTS_MIT = {
'bus_type': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
If no reference product clearly matches, respond with VALUE:UNKNOWN.
Step 1: Find the reference product that best matches the query product.
Step 2: Copy the bus_type value from that reference product exactly.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache - WD60EZAZ
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | bus_type: SATA III | model_number: WD60EZAZ
Best match: WD Blue 6TB WD60EZAZ -> bus_type: SATA III
VALUE:SATA III

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | bus_type: PCIe 3.0 x4 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> bus_type: PCIe 3.0 x4
VALUE:PCIe 3.0 x4

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> bus_type: [value from reference]
VALUE:""",

'model_number': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT generate or guess a model number.
Copy the EXACT model_number from the best matching reference product character by character.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

CRITICAL: model_numbers look like GV-N3080GAMING OC-10GD or CSSD-F960GBMP510.
Pay attention to every character -- GV-N166SOC-6GD and GV-N1660OC-6GD are DIFFERENT products.
If you are uncertain between two similar SKUs, respond with VALUE:UNKNOWN.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - SATA 6Gb/s 256MB Cache 3.5 Inch
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB Hard Drive -> model_number: WD60EZAZ
VALUE:WD60EZAZ

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD PCIe Gen3 x4 NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> model_number: CSSD-F960GBMP510
VALUE:CSSD-F960GBMP510

Example 3 (near-identical SKUs):
Query: Gigabyte GeForce GTX 1660 SUPER OC 6G graphics card
Reference products:
  - title: Gigabyte GTX 1660 Ti OC 6G | model_number: GV-N166TOC-6GD | brand: Gigabyte
  - title: Gigabyte GTX 1660 SUPER OC 6G | model_number: GV-N166SOC-6GD | brand: Gigabyte
Best match: GTX 1660 SUPER OC (not Ti) -> model_number: GV-N166SOC-6GD
VALUE:GV-N166SOC-6GD

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model_number: [exact value from reference]
VALUE:""",

'model': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
Copy the exact model name from the best matching reference product.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB -> model: WD Blue
VALUE:WD Blue

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 -> model: Force Series MP510
VALUE:Force Series MP510

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model: [value from reference]
VALUE:""",

'read_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact read_speed_mb_s number.
Return a number only. Do NOT return the write speed.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> read_speed_mb_s: 3480
VALUE:3480

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> read_speed_mb_s: [value from reference]
VALUE:""",

'write_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact write_speed_mb_s number.
Return a number only. Do NOT return the read speed.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> write_speed_mb_s: 3000
VALUE:3000

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> write_speed_mb_s: [value from reference]
VALUE:""",

'height_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact height_mm number.
Return a number only. Do NOT confuse height with width or length.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> height_mm: 46
VALUE:46

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> height_mm: [value from reference]
VALUE:""",

'width_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact width_mm number.
Return a number only. Do NOT confuse width with height or length.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> width_mm: 127
VALUE:127

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> width_mm: [value from reference]
VALUE:"""
}

print('Prompts ready.')

def row_to_text_bge_doc(row):
    td = row.get('title_description','')
    if pd.notna(td) and str(td).strip(): return str(td).strip()[:400]
    attrs = ['title','model','model_number','brand','product_type']
    text  = ' | '.join([str(row[a]) for a in attrs if pd.notna(row.get(a))])
    desc  = row.get('description','')
    if pd.notna(desc) and str(desc).strip(): text += ' | ' + str(desc).strip()[:200]
    return text

def format_candidates(candidates):
    lines = []
    for _, c in candidates.iterrows():
        fields = {k: str(v) for k,v in c.items()
                  if k not in SKIP_FIELDS and pd.notna(v)
                  and str(v).strip().lower() not in {'','nan','none'}}
        if fields:
            lines.append('  - ' + ' | '.join(f'{k}: {v}' for k,v in fields.items()))
    return '\n'.join(lines) if lines else '  (no candidates retrieved)'

def retrieve_top_n(q_emb, kb_embs, n):
    scores  = util.cos_sim(q_emb, kb_embs)[0]
    top_idx = np.argsort(-scores.cpu().numpy())[:n]
    return top_idx

def rerank(query_text, top_n_idx, kb_texts, k):
    cands_texts = [kb_texts[i] for i in top_n_idx]
    pairs       = [[query_text[:300], t[:300]] for t in cands_texts]
    scores      = cross_encoder.predict(pairs)
    top_k_local = np.argsort(-scores)[:k]
    return top_n_idx[top_k_local], scores[top_k_local]

def predict_with_timeout(predict_fn, timeout=60):
    result = ['UNKNOWN']
    def target():
        try: result[0] = predict_fn()
        except Exception as e:
            print(f'    Error: {e}')
            result[0] = 'UNKNOWN'
    t = threading.Thread(target=target)
    t.start(); t.join(timeout=timeout)
    if t.is_alive(): print('  WARNING TIMEOUT'); return 'UNKNOWN'
    return result[0]

kb_texts_bge = kb.apply(row_to_text_bge_doc, axis=1).tolist()
print('Helper functions ready.')


OPENAI_KEY = os.environ.get('OPENAI_API_KEY', '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')

RUNS = [
    {
        'model_name':   'gpt-4o-mini',
        'results_dir':  'results_ohne_UNKNOWN',
        'exp_file':     'results_ohne_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
        'ckpt_file':    'results_ohne_UNKNOWN/exp6_bge_reranker_gpt4omini_checkpoint.csv',
        'prompts':      PROMPTS_OHNE,
        'config_label': 'Exp 6: BGE+RR + gpt-4o-mini (ohne UNKNOWN)',
        'config_col':   'RAG-BGE-RR-GPT4oMini-ohne',
    },
    {
        'model_name':   'gpt-4o-mini',
        'results_dir':  'results_mit_UNKNOWN',
        'exp_file':     'results_mit_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
        'ckpt_file':    'results_mit_UNKNOWN/exp6_bge_reranker_gpt4omini_checkpoint.csv',
        'prompts':      PROMPTS_MIT,
        'config_label': 'Exp 6: BGE+RR + gpt-4o-mini (mit UNKNOWN)',
        'config_col':   'RAG-BGE-RR-GPT4oMini-mit',
    },
    {
        'model_name':   'gpt-5.4-mini',
        'results_dir':  'results_ohne_UNKNOWN',
        'exp_file':     'results_ohne_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
        'ckpt_file':    'results_ohne_UNKNOWN/exp6_bge_reranker_gpt54mini_checkpoint.csv',
        'prompts':      PROMPTS_OHNE,
        'config_label': 'Exp 6: BGE+RR + gpt-5.4-mini (ohne UNKNOWN)',
        'config_col':   'RAG-BGE-RR-GPT54Mini-ohne',
    },
    {
        'model_name':   'gpt-5.4-mini',
        'results_dir':  'results_mit_UNKNOWN',
        'exp_file':     'results_mit_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
        'ckpt_file':    'results_mit_UNKNOWN/exp6_bge_reranker_gpt54mini_checkpoint.csv',
        'prompts':      PROMPTS_MIT,
        'config_label': 'Exp 6: BGE+RR + gpt-5.4-mini (mit UNKNOWN)',
        'config_col':   'RAG-BGE-RR-GPT54Mini-mit',
    },
]

exp_dfs = {}

for run in RUNS:
    os.makedirs(run['results_dir'], exist_ok=True)
    print(f'\n{"="*65}')
    print(f'RUN: {run["config_label"]}')
    print(f'{"="*65}')

    predict_model = ChatOpenAI(model=run['model_name'], temperature=0, api_key=OPENAI_KEY)
    test = predict_model.invoke([HumanMessage(content='Say OK')])
    print(f'{run["model_name"]} OK: {repr(test.content[:20])}')

    if os.path.exists(run['exp_file']):
        print(f'Results already exist -- loading {run["exp_file"]}')
        df = pd.read_csv(run['exp_file'])
        exp_dfs[run['config_label']] = evaluate_and_save(df, run['config_label'], run['exp_file'])
        continue

    if os.path.exists(run['ckpt_file']) and os.path.getsize(run['ckpt_file']) > 0:
        try:
            checkpoint_df = pd.read_csv(run['ckpt_file'])
            if len(checkpoint_df) > 0:
                predictions     = checkpoint_df.to_dict('records')
                completed_tasks = set(zip(checkpoint_df['df1_idx'], checkpoint_df['attribute']))
                print(f'Resuming from checkpoint: {len(predictions)}/{len(eval_df)} done.')
            else:
                raise ValueError('Empty checkpoint')
        except Exception as e:
            print(f'Checkpoint unreadable ({e}) -- starting fresh.')
            predictions, completed_tasks = [], set()
    else:
        predictions, completed_tasks = [], set()
        print('Starting fresh.')

    _prompts = run['prompts']
    t0 = time.time()

    for i, (_, task) in enumerate(eval_df.iterrows()):
        idx, attr, gt = task['df1_idx'], task['attribute'], task['ground_truth']
        if (idx, attr) in completed_tasks:
            print(f'  [{i+1}/{len(eval_df)}] Skipping Row {idx} | {attr} -- already done')
            continue

        text       = query_df.loc[idx, 'title_description']
        q_emb      = bge_query_embs[query_idx_to_pos[idx]]
        query_text = row_to_text_bge_doc(query_df.loc[idx])

        top_n_idx    = retrieve_top_n(q_emb, bge_kb_embs, TOP_N)
        top_k_idx, _ = rerank(query_text, top_n_idx, kb_texts_bge, TOP_K)
        candidates   = kb.iloc[top_k_idx]

        def predict(_p=_prompts, _a=attr, _t=text, _c=candidates):
            prompt = _p[_a].format(text=str(_t)[:500], candidates=format_candidates(_c))
            return parse_response(predict_model.invoke([HumanMessage(content=prompt)]).content.strip(), _a)

        predicted = predict_with_timeout(predict, timeout=60)
        eta = (time.time()-t0) / (len(predictions)+1) * (len(eval_df)-i-1)
        print(f'  [{i+1}/{len(eval_df)}] Row {idx} | {attr:<22} | '
              f'GT: {str(gt):<25} | Pred: {predicted:<25} | ETA: {eta/60:.1f}min')

        predictions.append({
            'df1_idx':      idx,
            'config':       run['config_col'],
            'attribute':    attr,
            'is_numeric':   task['is_numeric'],
            'ground_truth': gt,
            'predicted':    predicted,
            'unknown':      predicted == 'UNKNOWN',
        })
        completed_tasks.add((idx, attr))

        if len(predictions) % CHECKPOINT_EVERY == 0:
            pd.DataFrame(predictions).to_csv(run['ckpt_file'], index=False)
            print(f'  + Checkpoint saved ({len(predictions)}/{len(eval_df)} done)')

    pd.DataFrame(predictions).to_csv(run['ckpt_file'], index=False)
    result_df = evaluate_and_save(pd.DataFrame(predictions), run['config_label'], run['exp_file'])
    exp_dfs[run['config_label']] = result_df
    if os.path.exists(run['ckpt_file']):
        os.remove(run['ckpt_file'])
    print(f'Done in {time.time()-t0:.1f}s')

print('\nAll BGE runs complete.')


# # TE

# In[ ]:


import sys
sys.path.insert(0, '/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI/venv/lib64/python3.12/site-packages')
sys.path.insert(0, '/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI/venv/lib/python3.12/site-packages')
sys.path.append('/home/ma/ma_ma/ma_mpandya/RAG_Data_Cleaning/PyDI')

import os, re, glob, time, random, threading
import torch
import numpy as np
import pandas as pd
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage
from sentence_transformers import CrossEncoder, util

random.seed(42); np.random.seed(42); torch.manual_seed(42)

print(f'PyTorch: {torch.__version__}')
print(f'CUDA: {torch.cuda.is_available()} -- {torch.cuda.get_device_name(0) if torch.cuda.is_available() else "CPU"}')

TARGET_ATTRIBUTES  = ['bus_type','model_number','model','read_speed_mb_s','write_speed_mb_s','height_mm','width_mm']
NUMERIC_ATTRIBUTES = {'read_speed_mb_s','write_speed_mb_s','height_mm','width_mm'}
TEXT_ATTRIBUTES    = {'bus_type','model_number','model'}
HF_CACHE           = '/home/ma/ma_ma/ma_mpandya/.cache/huggingface/hub'
EMBEDDINGS_DIR     = 'embeddings'
SKIP_FIELDS        = {'id','url','description','title_description','price','priceCurrency','cluster_id'}
TOP_N              = 20
TOP_K              = 5
CHECKPOINT_EVERY   = 5

DATA_DIR = 'normalized_products'
df1      = pd.read_json(f'{DATA_DIR}/dataset_1_normalized.json')
df2      = pd.read_json(f'{DATA_DIR}/dataset_2_normalized.json')
df3      = pd.read_json(f'{DATA_DIR}/dataset_3_normalized.json')
df4      = pd.read_json(f'{DATA_DIR}/dataset_4_normalized.json')
kb_full  = pd.concat([df2, df3, df4], ignore_index=True)
kb       = kb_full.copy()

assert os.path.exists('eval_set.csv'), 'Run exp_setup.ipynb first'
eval_df          = pd.read_csv('eval_set.csv')
query_indices    = pd.read_csv('query_indices.csv').iloc[:,0].tolist()
query_df         = df1.loc[query_indices].copy()
query_idx_to_pos = {idx: pos for pos, idx in enumerate(query_df.index)}

print(f'KB: {len(kb):,} rows | Query rows: {len(query_df)} | Eval tasks: {len(eval_df)}')

print('Loading TE (OpenAI) embeddings...')
bge_kb_embs    = torch.load(f'{EMBEDDINGS_DIR}/openai_kb.pt',    map_location='cpu')
bge_query_embs = torch.load(f'{EMBEDDINGS_DIR}/openai_query.pt', map_location='cpu')
print(f'  TE KB: {bge_kb_embs.shape} | Query: {bge_query_embs.shape}')

CE_SNAP = glob.glob(f'{HF_CACHE}/models--cross-encoder--ms-marco-MiniLM-L-6-v2/snapshots/*/')
CE_PATH = CE_SNAP[0].rstrip('/') if CE_SNAP else 'cross-encoder/ms-marco-MiniLM-L-6-v2'
cross_encoder = CrossEncoder(CE_PATH)
print(f'CrossEncoder: {CE_PATH}')

def is_correct_standard(predicted, ground_truth, attribute):
    if not predicted or str(predicted).strip().lower() in {'','nan','none','unknown','null'}:
        return False
    if attribute in NUMERIC_ATTRIBUTES:
        try:
            p = float(str(predicted).replace(',','').strip())
            g = float(str(ground_truth).replace(',','').strip())
            return p == g
        except: return False
    p, g = str(predicted).lower().strip(), str(ground_truth).lower().strip()
    return p == g or p in g or g in p

def evaluate_ce(predicted, ground_truth, attribute):
    if predicted == 'UNKNOWN' or str(predicted).lower() in {'nan','none','null',''}:
        return 'wrong'
    if attribute in NUMERIC_ATTRIBUTES:
        try:
            p = float(str(predicted).replace(',','').strip())
            g = float(str(ground_truth).replace(',','').strip())
            if g == 0: return 'correct' if p == 0 else 'wrong'
            r = abs(p-g)/abs(g)
            return 'correct' if r == 0.0 else ('acceptable' if r <= 0.10 else 'wrong')
        except: return 'wrong'
    score = cross_encoder.predict([[ground_truth, predicted]])[0]
    return 'correct' if score > 2.0 else ('acceptable' if score > -1.0 else 'wrong')

def parse_response(text, attribute):
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith('VALUE:'):
            val = line.split(':',1)[1].strip().strip('"').strip("'")
            return 'UNKNOWN' if val.upper() in {'UNKNOWN','NONE','NAN','NULL',''} else val
    pat = rf'{attribute}\s*[:\u2192>\-]+\s*([^\s|]+)'
    m = re.search(pat, text, re.IGNORECASE)
    if m:
        val = m.group(1).strip().strip('"').strip("'").strip('[]')
        if val.upper() != 'UNKNOWN' and len(val)>2 and val.lower() not in {'none','nan','null','exact','value'}:
            return val
    if attribute in NUMERIC_ATTRIBUTES:
        nums = re.findall(r'\b\d+\.?\d*\b', text)
        if nums: return nums[0]
    cleaned = text.strip().strip('"').strip("'")
    if cleaned.upper().startswith('VALUE:'): cleaned = cleaned.split(':',1)[1].strip()
    if cleaned and len(cleaned)<80 and '\n' not in cleaned and cleaned.upper() not in {'UNKNOWN','NONE','NULL','NAN',''}:
        return cleaned
    return 'UNKNOWN'

def fix_prediction(pred):
    if isinstance(pred, str) and pred.strip().upper().startswith('VALUE:'):
        val = pred.strip().split(':',1)[1].strip()
        return 'UNKNOWN' if val.upper() in {'UNKNOWN','NONE','NULL','NAN',''} else val
    return pred

def evaluate_and_save(results_df, config_name, filepath):
    results_df['predicted']        = results_df['predicted'].apply(fix_prediction)
    results_df['unknown']          = results_df['predicted'] == 'UNKNOWN'
    results_df['correct_standard'] = results_df.apply(
        lambda r: is_correct_standard(r['predicted'], r['ground_truth'], r['attribute']), axis=1)
    results_df['ce_judgment']      = [
        evaluate_ce(r['predicted'], r['ground_truth'], r['attribute'])
        for _, r in results_df.iterrows()]
    results_df.to_csv(filepath, index=False)
    std = results_df['correct_standard'].mean()
    ce  = results_df['ce_judgment'].isin(['correct','acceptable']).mean()
    unk = results_df['unknown'].mean()
    print(f'\n{"="*60}\nRESULTS -- {config_name}\n{"="*60}')
    print(f'Standard accuracy:    {std:.3f} ({std*100:.1f}%)')
    print(f'CE eval (c+a):        {ce:.3f} ({ce*100:.1f}%)')
    print(f'UNKNOWN rate:         {unk:.3f} ({unk*100:.1f}%)')
    print(f'Total tasks:          {len(results_df)}')
    print('\nPer-attribute:')
    print(results_df.groupby('attribute').agg(
        n=('correct_standard','count'),
        std_acc=('correct_standard','mean'),
        ce_acc=('ce_judgment', lambda x: x.isin(['correct','acceptable']).mean()),
        unknown=('unknown','mean')
    ).round(3).to_string())
    print(f'\n+ Saved to {filepath}')
    return results_df

print('Eval functions ready.')

PROMPTS_OHNE = {
'bus_type': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
You MUST always provide a value. Pick the best matching reference product even if uncertain.
Step 1: Find the reference product that best matches the query product.
Step 2: Copy the bus_type value from that reference product exactly.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache - WD60EZAZ
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | bus_type: SATA III | model_number: WD60EZAZ
Best match: WD Blue 6TB WD60EZAZ -> bus_type: SATA III
VALUE:SATA III

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | bus_type: PCIe 3.0 x4 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> bus_type: PCIe 3.0 x4
VALUE:PCIe 3.0 x4

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> bus_type: [value from reference]
VALUE:""",

'model_number': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT generate or guess a model number.
Copy the EXACT model_number from the best matching reference product character by character.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

CRITICAL: model_numbers look like GV-N3080GAMING OC-10GD or CSSD-F960GBMP510.
Pay attention to every character -- GV-N166SOC-6GD and GV-N1660OC-6GD are DIFFERENT products.
If uncertain between similar SKUs, pick the one whose title most closely matches the query.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - SATA 6Gb/s 256MB Cache 3.5 Inch
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB Hard Drive -> model_number: WD60EZAZ
VALUE:WD60EZAZ

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD PCIe Gen3 x4 NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> model_number: CSSD-F960GBMP510
VALUE:CSSD-F960GBMP510

Example 3 (near-identical SKUs):
Query: Gigabyte GeForce GTX 1660 SUPER OC 6G graphics card
Reference products:
  - title: Gigabyte GTX 1660 Ti OC 6G | model_number: GV-N166TOC-6GD | brand: Gigabyte
  - title: Gigabyte GTX 1660 SUPER OC 6G | model_number: GV-N166SOC-6GD | brand: Gigabyte
Best match: GTX 1660 SUPER OC (not Ti) -> model_number: GV-N166SOC-6GD
VALUE:GV-N166SOC-6GD

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model_number: [exact value from reference]
VALUE:""",

'model': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
Copy the exact model name from the best matching reference product.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB -> model: WD Blue
VALUE:WD Blue

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 -> model: Force Series MP510
VALUE:Force Series MP510

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model: [value from reference]
VALUE:""",

'read_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact read_speed_mb_s number.
Return a number only. Do NOT return the write speed.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> read_speed_mb_s: 3480
VALUE:3480

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> read_speed_mb_s: [value from reference]
VALUE:""",

'write_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact write_speed_mb_s number.
Return a number only. Do NOT return the read speed.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> write_speed_mb_s: 3000
VALUE:3000

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> write_speed_mb_s: [value from reference]
VALUE:""",

'height_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact height_mm number.
Return a number only. Do NOT confuse height with width or length.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> height_mm: 46
VALUE:46

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> height_mm: [value from reference]
VALUE:""",

'width_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact width_mm number.
Return a number only. Do NOT confuse width with height or length.
You MUST always provide a value. Pick the best matching reference product even if uncertain.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> width_mm: 127
VALUE:127

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> width_mm: [value from reference]
VALUE:"""
}

PROMPTS_MIT = {
'bus_type': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
If no reference product clearly matches, respond with VALUE:UNKNOWN.
Step 1: Find the reference product that best matches the query product.
Step 2: Copy the bus_type value from that reference product exactly.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache - WD60EZAZ
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | bus_type: SATA III | model_number: WD60EZAZ
Best match: WD Blue 6TB WD60EZAZ -> bus_type: SATA III
VALUE:SATA III

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | bus_type: PCIe 3.0 x4 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> bus_type: PCIe 3.0 x4
VALUE:PCIe 3.0 x4

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> bus_type: [value from reference]
VALUE:""",

'model_number': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT generate or guess a model number.
Copy the EXACT model_number from the best matching reference product character by character.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

CRITICAL: model_numbers look like GV-N3080GAMING OC-10GD or CSSD-F960GBMP510.
Pay attention to every character -- GV-N166SOC-6GD and GV-N1660OC-6GD are DIFFERENT products.
If you are uncertain between two similar SKUs, respond with VALUE:UNKNOWN.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - SATA 6Gb/s 256MB Cache 3.5 Inch
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB Hard Drive -> model_number: WD60EZAZ
VALUE:WD60EZAZ

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD PCIe Gen3 x4 NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 960GB -> model_number: CSSD-F960GBMP510
VALUE:CSSD-F960GBMP510

Example 3 (near-identical SKUs):
Query: Gigabyte GeForce GTX 1660 SUPER OC 6G graphics card
Reference products:
  - title: Gigabyte GTX 1660 Ti OC 6G | model_number: GV-N166TOC-6GD | brand: Gigabyte
  - title: Gigabyte GTX 1660 SUPER OC 6G | model_number: GV-N166SOC-6GD | brand: Gigabyte
Best match: GTX 1660 SUPER OC (not Ti) -> model_number: GV-N166SOC-6GD
VALUE:GV-N166SOC-6GD

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model_number: [exact value from reference]
VALUE:""",

'model': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Do NOT use your own knowledge.
Copy the exact model name from the best matching reference product.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: WD Blue 6TB Desktop Hard Disk Drive - 5400 RPM SATA 6Gb/s 256MB Cache
Reference products:
  - title: WD Blue 6TB Hard Drive WD60EZAZ | brand: Western Digital | model: WD Blue | model_number: WD60EZAZ
Best match: WD Blue 6TB -> model: WD Blue
VALUE:WD Blue

Example 2:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe
Reference products:
  - title: Corsair Force MP510 960GB NVMe SSD | brand: Corsair | model: Force Series MP510 | model_number: CSSD-F960GBMP510
Best match: Force MP510 -> model: Force Series MP510
VALUE:Force Series MP510

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> model: [value from reference]
VALUE:""",

'read_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact read_speed_mb_s number.
Return a number only. Do NOT return the write speed.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> read_speed_mb_s: 3480
VALUE:3480

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> read_speed_mb_s: [value from reference]
VALUE:""",

'write_speed_mb_s': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact write_speed_mb_s number.
Return a number only. Do NOT return the read speed.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: CORSAIR Force Series MP510 960GB M.2 SSD NVMe PCIe Gen3
Reference products:
  - title: Corsair Force MP510 960GB NVMe | model_number: CSSD-F960GBMP510 | read_speed_mb_s: 3480 | write_speed_mb_s: 3000
Best match: Force MP510 960GB -> write_speed_mb_s: 3000
VALUE:3000

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> write_speed_mb_s: [value from reference]
VALUE:""",

'height_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact height_mm number.
Return a number only. Do NOT confuse height with width or length.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> height_mm: 46
VALUE:46

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> height_mm: [value from reference]
VALUE:""",

'width_mm': """\
You are a product data expert filling missing values in a product database.
You MUST use ONLY the reference products below. Copy the exact width_mm number.
Return a number only. Do NOT confuse width with height or length.
If no reference product clearly matches, respond with VALUE:UNKNOWN.

Example 1:
Query: MSI GeForce GTX 1660 Ti GAMING X 6G graphics card
Reference products:
  - title: MSI GTX 1660 Ti GAMING X 6G | model_number: V375-040R | height_mm: 46 | width_mm: 127 | length_mm: 247
Best match: GTX 1660 Ti GAMING X -> width_mm: 127
VALUE:127

Now fill the missing value:
Query: {text}
Reference products:
{candidates}
Best match: [identify matching product] -> width_mm: [value from reference]
VALUE:"""
}

print('Prompts ready.')

def row_to_text_bge_doc(row):
    td = row.get('title_description','')
    if pd.notna(td) and str(td).strip(): return str(td).strip()[:400]
    attrs = ['title','model','model_number','brand','product_type']
    text  = ' | '.join([str(row[a]) for a in attrs if pd.notna(row.get(a))])
    desc  = row.get('description','')
    if pd.notna(desc) and str(desc).strip(): text += ' | ' + str(desc).strip()[:200]
    return text

def format_candidates(candidates):
    lines = []
    for _, c in candidates.iterrows():
        fields = {k: str(v) for k,v in c.items()
                  if k not in SKIP_FIELDS and pd.notna(v)
                  and str(v).strip().lower() not in {'','nan','none'}}
        if fields:
            lines.append('  - ' + ' | '.join(f'{k}: {v}' for k,v in fields.items()))
    return '\n'.join(lines) if lines else '  (no candidates retrieved)'

def retrieve_top_n(q_emb, kb_embs, n):
    scores  = util.cos_sim(q_emb, kb_embs)[0]
    top_idx = np.argsort(-scores.cpu().numpy())[:n]
    return top_idx

def rerank(query_text, top_n_idx, kb_texts, k):
    cands_texts = [kb_texts[i] for i in top_n_idx]
    pairs       = [[query_text[:300], t[:300]] for t in cands_texts]
    scores      = cross_encoder.predict(pairs)
    top_k_local = np.argsort(-scores)[:k]
    return top_n_idx[top_k_local], scores[top_k_local]

def predict_with_timeout(predict_fn, timeout=60):
    result = ['UNKNOWN']
    def target():
        try: result[0] = predict_fn()
        except Exception as e:
            print(f'    Error: {e}')
            result[0] = 'UNKNOWN'
    t = threading.Thread(target=target)
    t.start(); t.join(timeout=timeout)
    if t.is_alive(): print('  WARNING TIMEOUT'); return 'UNKNOWN'
    return result[0]

kb_texts_bge = kb.apply(row_to_text_bge_doc, axis=1).tolist()
print('Helper functions ready.')


OPENAI_KEY = os.environ.get('OPENAI_API_KEY', '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')

RUNS = [
    {
        'model_name':   'gpt-4o-mini',
        'results_dir':  'results_ohne_UNKNOWN',
        'exp_file':     'results_ohne_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
        'ckpt_file':    'results_ohne_UNKNOWN/exp7_te_reranker_gpt4omini_checkpoint.csv',
        'prompts':      PROMPTS_OHNE,
        'config_label': 'Exp 7: TE+RR + gpt-4o-mini (ohne UNKNOWN)',
        'config_col':   'RAG-TE-RR-GPT4oMini-ohne',
    },
    {
        'model_name':   'gpt-4o-mini',
        'results_dir':  'results_mit_UNKNOWN',
        'exp_file':     'results_mit_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
        'ckpt_file':    'results_mit_UNKNOWN/exp7_te_reranker_gpt4omini_checkpoint.csv',
        'prompts':      PROMPTS_MIT,
        'config_label': 'Exp 7: TE+RR + gpt-4o-mini (mit UNKNOWN)',
        'config_col':   'RAG-TE-RR-GPT4oMini-mit',
    },
    {
        'model_name':   'gpt-5.4-mini',
        'results_dir':  'results_ohne_UNKNOWN',
        'exp_file':     'results_ohne_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
        'ckpt_file':    'results_ohne_UNKNOWN/exp7_te_reranker_gpt54mini_checkpoint.csv',
        'prompts':      PROMPTS_OHNE,
        'config_label': 'Exp 7: TE+RR + gpt-5.4-mini (ohne UNKNOWN)',
        'config_col':   'RAG-TE-RR-GPT54Mini-ohne',
    },
    {
        'model_name':   'gpt-5.4-mini',
        'results_dir':  'results_mit_UNKNOWN',
        'exp_file':     'results_mit_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
        'ckpt_file':    'results_mit_UNKNOWN/exp7_te_reranker_gpt54mini_checkpoint.csv',
        'prompts':      PROMPTS_MIT,
        'config_label': 'Exp 7: TE+RR + gpt-5.4-mini (mit UNKNOWN)',
        'config_col':   'RAG-TE-RR-GPT54Mini-mit',
    },
]

exp_dfs = {}

for run in RUNS:
    os.makedirs(run['results_dir'], exist_ok=True)
    print(f'\n{"="*65}')
    print(f'RUN: {run["config_label"]}')
    print(f'{"="*65}')

    predict_model = ChatOpenAI(model=run['model_name'], temperature=0, api_key=OPENAI_KEY)
    test = predict_model.invoke([HumanMessage(content='Say OK')])
    print(f'{run["model_name"]} OK: {repr(test.content[:20])}')

    if os.path.exists(run['exp_file']):
        print(f'Results already exist -- loading {run["exp_file"]}')
        df = pd.read_csv(run['exp_file'])
        exp_dfs[run['config_label']] = evaluate_and_save(df, run['config_label'], run['exp_file'])
        continue

    if os.path.exists(run['ckpt_file']) and os.path.getsize(run['ckpt_file']) > 0:
        try:
            checkpoint_df = pd.read_csv(run['ckpt_file'])
            if len(checkpoint_df) > 0:
                predictions     = checkpoint_df.to_dict('records')
                completed_tasks = set(zip(checkpoint_df['df1_idx'], checkpoint_df['attribute']))
                print(f'Resuming from checkpoint: {len(predictions)}/{len(eval_df)} done.')
            else:
                raise ValueError('Empty checkpoint')
        except Exception as e:
            print(f'Checkpoint unreadable ({e}) -- starting fresh.')
            predictions, completed_tasks = [], set()
    else:
        predictions, completed_tasks = [], set()
        print('Starting fresh.')

    _prompts = run['prompts']
    t0 = time.time()

    for i, (_, task) in enumerate(eval_df.iterrows()):
        idx, attr, gt = task['df1_idx'], task['attribute'], task['ground_truth']
        if (idx, attr) in completed_tasks:
            print(f'  [{i+1}/{len(eval_df)}] Skipping Row {idx} | {attr} -- already done')
            continue

        text       = query_df.loc[idx, 'title_description']
        q_emb      = bge_query_embs[query_idx_to_pos[idx]]
        query_text = row_to_text_bge_doc(query_df.loc[idx])

        top_n_idx    = retrieve_top_n(q_emb, bge_kb_embs, TOP_N)
        top_k_idx, _ = rerank(query_text, top_n_idx, kb_texts_bge, TOP_K)
        candidates   = kb.iloc[top_k_idx]

        def predict(_p=_prompts, _a=attr, _t=text, _c=candidates):
            prompt = _p[_a].format(text=str(_t)[:500], candidates=format_candidates(_c))
            return parse_response(predict_model.invoke([HumanMessage(content=prompt)]).content.strip(), _a)

        predicted = predict_with_timeout(predict, timeout=60)
        eta = (time.time()-t0) / (len(predictions)+1) * (len(eval_df)-i-1)
        print(f'  [{i+1}/{len(eval_df)}] Row {idx} | {attr:<22} | '
              f'GT: {str(gt):<25} | Pred: {predicted:<25} | ETA: {eta/60:.1f}min')

        predictions.append({
            'df1_idx':      idx,
            'config':       run['config_col'],
            'attribute':    attr,
            'is_numeric':   task['is_numeric'],
            'ground_truth': gt,
            'predicted':    predicted,
            'unknown':      predicted == 'UNKNOWN',
        })
        completed_tasks.add((idx, attr))

        if len(predictions) % CHECKPOINT_EVERY == 0:
            pd.DataFrame(predictions).to_csv(run['ckpt_file'], index=False)
            print(f'  + Checkpoint saved ({len(predictions)}/{len(eval_df)} done)')

    pd.DataFrame(predictions).to_csv(run['ckpt_file'], index=False)
    result_df = evaluate_and_save(pd.DataFrame(predictions), run['config_label'], run['exp_file'])
    exp_dfs[run['config_label']] = result_df
    if os.path.exists(run['ckpt_file']):
        os.remove(run['ckpt_file'])
    print(f'Done in {time.time()-t0:.1f}s')

print('\nAll TE runs complete.')


# # Error Analysis:

# In[ ]:


import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import numpy as np
import os

NUMERIC = {'read_speed_mb_s','write_speed_mb_s','height_mm','width_mm'}
os.makedirs('figures', exist_ok=True)

def fix(p):
    if isinstance(p,str) and p.strip().upper().startswith('VALUE:'):
        v = p.strip().split(':',1)[1].strip()
        return 'UNKNOWN' if v.upper() in {'UNKNOWN','NONE','NULL','NAN',''} else v
    return p

def is_unk(v): return str(v).strip().lower() in {'unknown','nan','none','null',''}

def correct(pred, gt, attr):
    if is_unk(pred): return False
    if attr in NUMERIC:
        try: return float(str(pred).replace(',','')) == float(str(gt).replace(',',''))
        except: return False
    p, g = str(pred).lower().strip(), str(gt).lower().strip()
    return p == g or p in g or g in p

FILE_MAP = {
    'BGE g4omini ohne': 'results_ohne_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
    'BGE g4omini mit':  'results_mit_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
    'BGE g54mini ohne': 'results_ohne_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
    'BGE g54mini mit':  'results_mit_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
    'TE g4omini ohne':  'results_ohne_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
    'TE g4omini mit':   'results_mit_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
    'TE g54mini ohne':  'results_ohne_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
    'TE g54mini mit':   'results_mit_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
}

dfs = {}
for label, path in FILE_MAP.items():
    try:
        df = pd.read_csv(path)
        df['predicted'] = df['predicted'].apply(fix)
        df['correct']   = df.apply(lambda r: correct(r['predicted'], r['ground_truth'], r['attribute']), axis=1)
        df['unk']       = df['predicted'].apply(is_unk)
        dfs[label] = df
        print(f'+ {label}')
    except FileNotFoundError:
        print(f'x NOT FOUND: {path}')

labels   = list(FILE_MAP.keys())
acc_vals = [dfs[l]['correct'].mean()*100 if l in dfs else 0 for l in labels]
unk_vals = [dfs[l]['unk'].mean()*100     if l in dfs else 0 for l in labels]

COLORS = ['#1E3A5F','#2563EB','#0891B2','#38BDF8','#065F46','#059669','#0D9488','#34D399']
C_UNK  = '#EF4444'

fig, axes = plt.subplots(1, 2, figsize=(20, 6))
fig.patch.set_facecolor('#FAFBFC')

for ax, vals, ylabel, title, use_unk_color in [
    (axes[0], acc_vals, 'Accuracy (%)',     'Standard Accuracy',  False),
    (axes[1], unk_vals, 'UNKNOWN Rate (%)', 'UNKNOWN Rate',       True),
]:
    ax.set_facecolor('#F8FAFC')
    for i, v in enumerate(vals):
        color = C_UNK if use_unk_color else COLORS[i]
        ax.bar(i, v, 0.6, color=color, alpha=0.88, edgecolor='white', linewidth=0.8, zorder=3)
        ax.text(i, v+0.5, f'{v:.1f}%', ha='center', va='bottom', fontsize=8.5, fontweight='bold',
                color=C_UNK if use_unk_color else 'black')
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=9)
    ax.set_ylim(0, max(vals or [1])*1.18 + 3)
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax.set_ylabel(ylabel)
    ax.set_title(title, fontsize=12, fontweight='bold')
    ax.grid(axis='y', alpha=0.25)
    ax.spines[['top','right']].set_visible(False)
    ax.axvline(3.5, color='#CBD5E1', lw=1.2, linestyle='--')
    ax.text(1.5, ax.get_ylim()[1]*0.96, 'BGE-large', ha='center', fontsize=9, color='#64748B', style='italic')
    ax.text(5.5, ax.get_ylim()[1]*0.96, 'TE-3-large', ha='center', fontsize=9, color='#64748B', style='italic')

plt.suptitle('All 8 Configurations -- Accuracy & UNKNOWN Rate', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig('figures/llm_all8_overview.png', dpi=150, bbox_inches='tight', facecolor='#FAFBFC')
plt.show()

bge_llama = pd.read_csv('results_ohne_UNKNOWN/exp4_rag_bge_reranker.csv')
te_llama  = pd.read_csv('results_ohne_UNKNOWN/exp5_rag_te_reranker.csv')
for df in [bge_llama, te_llama]:
    df['predicted'] = df['predicted'].apply(fix)
    df['correct']   = df.apply(lambda r: correct(r['predicted'], r['ground_truth'], r['attribute']), axis=1)

attrs = sorted(bge_llama['attribute'].unique())

panels = [
    ('BGE g4omini ohne', bge_llama, 'BGE GPT-4o-mini\nohne UNKNOWN'),
    ('BGE g4omini mit',  bge_llama, 'BGE GPT-4o-mini\nmit UNKNOWN'),
    ('BGE g54mini ohne', bge_llama, 'BGE GPT-5.4-mini\nohne UNKNOWN'),
    ('BGE g54mini mit',  bge_llama, 'BGE GPT-5.4-mini\nmit UNKNOWN'),
    ('TE g4omini ohne',  te_llama,  'TE GPT-4o-mini\nohne UNKNOWN'),
    ('TE g4omini mit',   te_llama,  'TE GPT-4o-mini\nmit UNKNOWN'),
    ('TE g54mini ohne',  te_llama,  'TE GPT-5.4-mini\nohne UNKNOWN'),
    ('TE g54mini mit',   te_llama,  'TE GPT-5.4-mini\nmit UNKNOWN'),
]

fig, axes = plt.subplots(2, 4, figsize=(24, 11))
fig.patch.set_facecolor('#FAFBFC')

for ax, (key, llama_df, title) in zip(axes.flatten(), panels):
    if key not in dfs:
        ax.set_title(f'{title}\n(missing)', fontsize=9); ax.axis('off'); continue
    model_df = dfs[key]
    deltas = [
        model_df[model_df['attribute']==a]['correct'].mean()*100
        - llama_df[llama_df['attribute']==a]['correct'].mean()*100
        for a in attrs
    ]
    colors = ['#059669' if d >= 0 else '#EF4444' for d in deltas]
    y = np.arange(len(attrs))
    ax.barh(y, deltas, color=colors, alpha=0.85, edgecolor='white')
    for i, v in enumerate(deltas):
        ax.text(v+(0.4 if v>=0 else -0.4), i, f'{v:+.1f}%',
                va='center', ha='left' if v>=0 else 'right', fontsize=7.5)
    ax.set_yticks(y); ax.set_yticklabels(attrs, fontsize=8.5)
    ax.axvline(0, color='#334155', lw=1.5)
    ax.set_title(title, fontsize=10, fontweight='bold')
    ax.grid(axis='x', alpha=0.25)
    ax.spines[['top','right']].set_visible(False)

plt.suptitle('Per-Attribute Accuracy Delta vs Llama 3.1 8B', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig('figures/llm_all8_delta_vs_llama.png', dpi=150, bbox_inches='tight', facecolor='#FAFBFC')
plt.show()

fig, axes = plt.subplots(2, 2, figsize=(18, 12))
fig.patch.set_facecolor('#FAFBFC')

model_panels = [
    (axes[0,0], 'BGE g4omini ohne', 'BGE g4omini mit',  bge_llama, 'BGE+RR GPT-4o-mini'),
    (axes[0,1], 'BGE g54mini ohne', 'BGE g54mini mit',  bge_llama, 'BGE+RR GPT-5.4-mini'),
    (axes[1,0], 'TE g4omini ohne',  'TE g4omini mit',   te_llama,  'TE+RR GPT-4o-mini'),
    (axes[1,1], 'TE g54mini ohne',  'TE g54mini mit',   te_llama,  'TE+RR GPT-5.4-mini'),
]

for ax, ohne_key, mit_key, llama_df, title in model_panels:
    missing = [k for k in [ohne_key, mit_key] if k not in dfs]
    if missing:
        ax.set_title(f'{title}\n(missing: {missing})', fontsize=10); ax.axis('off'); continue
    d_ohne = [dfs[ohne_key][dfs[ohne_key]['attribute']==a]['correct'].mean()*100
              - llama_df[llama_df['attribute']==a]['correct'].mean()*100 for a in attrs]
    d_mit  = [dfs[mit_key][dfs[mit_key]['attribute']==a]['correct'].mean()*100
              - llama_df[llama_df['attribute']==a]['correct'].mean()*100 for a in attrs]
    y = np.arange(len(attrs)); h = 0.36
    ax.barh(y-h/2, d_ohne, h, color='#1E3A5F', alpha=0.85, label='ohne UNKNOWN')
    ax.barh(y+h/2, d_mit,  h, color='#38BDF8', alpha=0.85, label='mit UNKNOWN')
    for i, v in enumerate(d_ohne):
        ax.text(v+(0.3 if v>=0 else -0.3), i-h/2, f'{v:+.1f}',
                va='center', ha='left' if v>=0 else 'right', fontsize=8)
    for i, v in enumerate(d_mit):
        ax.text(v+(0.3 if v>=0 else -0.3), i+h/2, f'{v:+.1f}',
                va='center', ha='left' if v>=0 else 'right', fontsize=8)
    ax.axvline(0, color='#334155', lw=1.5)
    ax.set_yticks(y); ax.set_yticklabels(attrs, fontsize=9)
    ax.set_xlabel('Accuracy Change vs Llama (pp)')
    ax.set_title(title, fontsize=12, fontweight='bold', color='#1E3A5F')
    ax.grid(axis='x', alpha=0.25)
    ax.spines[['top','right']].set_visible(False)
    ax.legend(fontsize=9, framealpha=0.9)

plt.suptitle('mit vs ohne UNKNOWN -- Per-Attribute Gain vs Llama 3.1 8B', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig('figures/llm_mit_vs_ohne_grouped.png', dpi=150, bbox_inches='tight', facecolor='#FAFBFC')
plt.show()


# In[ ]:


for key in ['BGE g4omini ohne','BGE g54mini ohne','TE g4omini ohne','TE g54mini ohne']:
    if key not in dfs: continue
    emb      = 'BGE' if 'BGE' in key else 'TE'
    llama_df = bge_llama if emb == 'BGE' else te_llama
    merged = llama_df[['df1_idx','attribute','ground_truth']].copy()
    merged['llama_pred']    = llama_df['predicted'].values
    merged['model_pred']    = dfs[key]['predicted'].values
    merged['llama_correct'] = llama_df['correct'].values
    merged['model_correct'] = dfs[key]['correct'].values
    merged['model_unk']     = dfs[key]['unk'].values
    interesting = merged[merged['llama_correct'] & ~merged['model_correct']]
    print(f'\n=== Llama correct, {key} wrong ({len(interesting)} cases) ===')
    print(f'{"Row":<6}{"Attr":<22}{"GT":<20}{"Llama":<25}{"Model":<25}')
    print('='*98)
    for _, r in interesting.iterrows():
        model_label = 'UNKNOWN x' if r['model_unk'] else str(r['model_pred']) + ' x'
        print(f'{int(r["df1_idx"]):<6}{r["attribute"]:<22}{str(r["ground_truth"]):<20}'
              f'{str(r["llama_pred"])+" ok":<25}{model_label:<25}')
    print(f'-> Model returned UNKNOWN: {interesting["model_unk"].sum()}')
    print(f'-> Model wrong value:      {(~interesting["model_correct"] & ~interesting["model_unk"]).sum()}')


# # Error Analysis (without visualization)

# In[ ]:


import pandas as pd
import numpy as np
from sentence_transformers import CrossEncoder
import glob, os

NUMERIC  = {'read_speed_mb_s','write_speed_mb_s','height_mm','width_mm'}
HF_CACHE = '/home/ma/ma_ma/ma_mpandya/.cache/huggingface/hub'
CE_SNAP  = glob.glob(f'{HF_CACHE}/models--cross-encoder--ms-marco-MiniLM-L-6-v2/snapshots/*/')
CE_PATH  = CE_SNAP[0].rstrip('/') if CE_SNAP else 'cross-encoder/ms-marco-MiniLM-L-6-v2'
cross_encoder = CrossEncoder(CE_PATH)

def fix(p):
    if isinstance(p,str) and p.strip().upper().startswith('VALUE:'):
        v = p.strip().split(':',1)[1].strip()
        return 'UNKNOWN' if v.upper() in {'UNKNOWN','NONE','NULL','NAN',''} else v
    return p

def is_unk(v): return str(v).strip().lower() in {'unknown','nan','none','null',''}

def std_correct(pred, gt, attr):
    if is_unk(pred): return False
    if attr in NUMERIC:
        try: return float(str(pred).replace(',','')) == float(str(gt).replace(',',''))
        except: return False
    p, g = str(pred).lower().strip(), str(gt).lower().strip()
    return p == g or p in g or g in p

def ce_correct(pred, gt, attr):
    if is_unk(pred): return False
    if attr in NUMERIC:
        try:
            p = float(str(pred).replace(',','').strip())
            g = float(str(gt).replace(',','').strip())
            if g == 0: return p == 0
            return abs(p-g)/abs(g) <= 0.10
        except: return False
    return cross_encoder.predict([[gt, pred]])[0] > -1.0

FILES = {
    'BGE Llama (ohne)':      'results_ohne_UNKNOWN/exp4_rag_bge_reranker.csv',
    'BGE g4omini (ohne)':    'results_ohne_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
    'BGE g54mini (ohne)':    'results_ohne_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
    'BGE Llama (mit)':       'results_mit_UNKNOWN/exp4_rag_bge_reranker.csv',
    'BGE g4omini (mit)':     'results_mit_UNKNOWN/exp6_bge_reranker_gpt4omini.csv',
    'BGE g54mini (mit)':     'results_mit_UNKNOWN/exp6_bge_reranker_gpt54mini.csv',
    'TE Llama (ohne)':       'results_ohne_UNKNOWN/exp5_rag_te_reranker.csv',
    'TE g4omini (ohne)':     'results_ohne_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
    'TE g54mini (ohne)':     'results_ohne_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
    'TE Llama (mit)':        'results_mit_UNKNOWN/exp5_rag_te_reranker.csv',
    'TE g4omini (mit)':      'results_mit_UNKNOWN/exp7_te_reranker_gpt4omini.csv',
    'TE g54mini (mit)':      'results_mit_UNKNOWN/exp7_te_reranker_gpt54mini.csv',
}

ATTRS = ['bus_type','model','model_number','read_speed_mb_s','write_speed_mb_s','height_mm','width_mm']
CONFIGS_OHNE = ['BGE Llama (ohne)','BGE g4omini (ohne)','BGE g54mini (ohne)',
                'TE Llama (ohne)', 'TE g4omini (ohne)', 'TE g54mini (ohne)']
CONFIGS_MIT  = ['BGE Llama (mit)', 'BGE g4omini (mit)', 'BGE g54mini (mit)',
                'TE Llama (mit)',  'TE g4omini (mit)',  'TE g54mini (mit)']

dfs_ea = {}
for label, path in FILES.items():
    try:
        df = pd.read_csv(path)
        df['predicted'] = df['predicted'].apply(fix)
        df['std'] = df.apply(lambda r: std_correct(r['predicted'], r['ground_truth'], r['attribute']), axis=1)
        df['ce']  = df.apply(lambda r: ce_correct(r['predicted'],  r['ground_truth'], r['attribute']), axis=1)
        df['unk'] = df['predicted'].apply(is_unk)
        dfs_ea[label] = df
        print(f'+ {label}')
    except FileNotFoundError:
        print(f'x NOT FOUND: {path}')

def build_table(config_list, metric):
    rows = []
    for attr in ATTRS:
        row = {'Attribute': attr, 'Type': 'Numeric' if attr in NUMERIC else 'Text'}
        for cfg in config_list:
            if cfg not in dfs_ea: row[cfg] = 'N/A'; continue
            sub = dfs_ea[cfg][dfs_ea[cfg]['attribute']==attr]
            row[cfg] = round(sub[metric].mean()*100, 1)
        rows.append(row)
    overall = {'Attribute': 'OVERALL', 'Type': ''}
    for cfg in config_list:
        if cfg not in dfs_ea: overall[cfg]='N/A'; continue
        overall[cfg] = round(dfs_ea[cfg][metric].mean()*100, 1)
    rows.append(overall)
    unk_row = {'Attribute': 'UNKNOWN rate', 'Type': ''}
    for cfg in config_list:
        if cfg not in dfs_ea: unk_row[cfg]='N/A'; continue
        unk_row[cfg] = round(dfs_ea[cfg]['unk'].mean()*100, 1)
    rows.append(unk_row)
    return pd.DataFrame(rows)

std_ohne = build_table(CONFIGS_OHNE, 'std')
ce_ohne  = build_table(CONFIGS_OHNE, 'ce')
std_mit  = build_table(CONFIGS_MIT,  'std')
ce_mit   = build_table(CONFIGS_MIT,  'ce')

for title, tbl in [('Std ohne UNKNOWN', std_ohne), ('CE ohne UNKNOWN', ce_ohne),
                   ('Std mit UNKNOWN',  std_mit),  ('CE mit UNKNOWN',  ce_mit)]:
    print(f'\n=== {title} ===')
    print(tbl.to_string(index=False))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY='1E3A5F'; TEAL='0891B2'; GREEN='059669'; RED='DC2626'; GOLD='D97706'
    WHITE='FFFFFF'; GREY='F1F5F9'; DKGRY='CBD5E1'

    def style_sheet(ws, df, title):
        ws.title = title[:31]
        ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
        ws['A1'] = title
        ws['A1'].font      = Font(bold=True, size=13, color=WHITE)
        ws['A1'].fill      = PatternFill('solid', fgColor=NAVY)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        thin = Side(style='thin', color=DKGRY)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_i, value=col)
            cell.font      = Font(bold=True, size=10, color=WHITE)
            cell.fill      = PatternFill('solid', fgColor=TEAL)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = border
        ws.row_dimensions[2].height = 42
        for row_i, (_, row) in enumerate(df.iterrows(), 3):
            is_special = str(row['Attribute']) in {'OVERALL','UNKNOWN rate'}
            is_numeric = row.get('Type','') == 'Numeric'
            for col_i, val in enumerate(row, 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.border    = border
                cell.alignment = Alignment(horizontal='center' if col_i>1 else 'left', vertical='center')
                if is_special:
                    cell.fill = PatternFill('solid', fgColor='E2E8F0')
                    cell.font = Font(bold=True, size=10)
                elif is_numeric:
                    cell.fill = PatternFill('solid', fgColor='EFF6FF')
                    cell.font = Font(size=10)
                elif row_i % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor=GREY)
                    cell.font = Font(size=10)
                else:
                    cell.font = Font(size=10)
                if col_i > 2 and isinstance(val, (int, float)):
                    if str(row['Attribute']) == 'UNKNOWN rate':
                        if val <= 5:    cell.font = Font(bold=True, size=10, color=GREEN)
                        elif val >= 20: cell.font = Font(bold=True, size=10, color=RED)
                        else:           cell.font = Font(bold=True, size=10, color=GOLD)
                    else:
                        if val >= 80:   cell.font = Font(bold=True, size=10, color=GREEN)
                        elif val >= 60: cell.font = Font(bold=True, size=10, color='0E7490')
                        elif val < 30:  cell.font = Font(bold=True, size=10, color=RED)
                        else:           cell.font = Font(size=10)
            ws.row_dimensions[row_i].height = 20
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        for col_i in range(3, len(df.columns)+1):
            ws.column_dimensions[get_column_letter(col_i)].width = 22

    os.makedirs('results', exist_ok=True)

    wb_ohne = openpyxl.Workbook()
    style_sheet(wb_ohne.active, std_ohne, 'Std Accuracy -- Ohne UNKNOWN')
    style_sheet(wb_ohne.create_sheet(), ce_ohne, 'CE Accuracy -- Ohne UNKNOWN')
    wb_ohne.save('results/llm_comparison_ohne_unknown.xlsx')
    print('\n+ results/llm_comparison_ohne_unknown.xlsx')

    wb_mit = openpyxl.Workbook()
    style_sheet(wb_mit.active, std_mit, 'Std Accuracy -- Mit UNKNOWN')
    style_sheet(wb_mit.create_sheet(), ce_mit, 'CE Accuracy -- Mit UNKNOWN')
    wb_mit.save('results/llm_comparison_mit_unknown.xlsx')
    print('+ results/llm_comparison_mit_unknown.xlsx')

except ImportError:
    print('openpyxl not found -- saving as CSV')
    std_ohne.to_csv('results/std_ohne_unknown.csv', index=False)
    ce_ohne.to_csv('results/ce_ohne_unknown.csv',   index=False)
    std_mit.to_csv('results/std_mit_unknown.csv',   index=False)
    ce_mit.to_csv('results/ce_mit_unknown.csv',     index=False)
    print('+ Saved 4 CSV files to results/')


# In[ ]:




